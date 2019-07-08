# -*- coding: utf-8 -*-
"""
Created on Wed Dec  5 15:57:09 2018

@author: VL58003
"""
# %%
import os
import openpyxl
import datetime
from openpyxl.styles.fonts import Font
from openpyxl.styles.fills import PatternFill
from tkinter import filedialog
from tkinter import Tk, messagebox
import re
import time

def makefolder():
    
    dir="C:"
    fld=filedialog.askdirectory(initialdir=dir) 
    os.chdir(fld)
    
    kyo=datetime.date.today()
    kyo_name="_{0:%Y%m%d}".format(kyo)
    
    list_root=[]
    list_file=[]
    
    for root,dirs,files in os.walk(fld):
        for file in files:
            list_root.append(root)
            list_file.append(files)
            list=dict(zip(list_root,list_file))
    
    wb=openpyxl.Workbook()
    ws=wb.active
    
    ws.title="フォルダ内ファイルまとめ"
    ws.column_dimensions["A"].width=100
    ws.column_dimensions["B"].width=60
    ws.column_dimensions["C"].width=10
    ws.column_dimensions["D"].width=5
    ws.column_dimensions["E"].width=20
    ws.column_dimensions["F"].width=20
    ws["A1"]="フォルダ名"
    ws["B1"]="ファイル名"
    ws["C1"]="ファイル種類"
    ws["D1"]="日付"
    ws["E1"]="作成日時"
    ws["F1"]="アクセス"
    
    greenyellow="ADFF2F"
    lightblue="87CEFA"
    lightsalmon="FFA07A"
    plum="DDA0DD"
    gold="FFD700"
    
    date=re.compile(r"\d{6}")
    
    length=1
    for i in list:
        fld2=i.replace("/","\\")
        ws["A"+str(length+1)]=i
        ws["A"+str(length+1)].hyperlink=i
        ws["A"+str(length+1)].font = Font(color="ff0000ff")
        for number,j in enumerate(list[i]):
            ws["B"+str(length+number+1)]=j
            ws["B"+str(length+number+1)].hyperlink=i+"\\"+j
            ws["E"+str(length+number+1)]=time.strftime("%Y/%m/%d %H:%M:%S",time.localtime(os.path.getctime(i+"\\"+j)))
            ws["F"+str(length+number+1)]=time.strftime("%Y/%m/%d %H:%M:%S",time.localtime(os.path.getatime(i+"\\"+j)))
            fn,ext=os.path.splitext(j)
            if ext==".xlsx" or ext==".xlsm" or ext==".XLSX":
                ws["B"+str(length+number+1)].fill=openpyxl.styles.PatternFill(patternType='solid',fgColor=greenyellow,bgColor=greenyellow)
                ws["C"+str(length+number+1)] = "Excel"
            elif ext==".docx":
                ws["B"+str(length+number+1)].fill=openpyxl.styles.PatternFill(patternType='solid',fgColor=lightblue,bgColor=lightblue)
                ws["C"+str(length+number+1)] = "Word"
            elif ext==".pptx":
                ws["B"+str(length+number+1)].fill=openpyxl.styles.PatternFill(patternType='solid',fgColor=lightsalmon,bgColor=lightsalmon)
                ws["C"+str(length+number+1)] = "PowerPoint"
            elif ext==".pdf":
                ws["B"+str(length+number+1)].fill=openpyxl.styles.PatternFill(patternType='solid',fgColor=plum,bgColor=plum)
                ws["C"+str(length+number+1)] = "PDF"
            elif ext==".jpg" or ext==".JPG" or ext==".png" or ext=="PNG":
                ws["B"+str(length+number+1)].fill=openpyxl.styles.PatternFill(patternType='solid',fgColor=gold,bgColor=gold)
                ws["C"+str(length+number+1)] = "Image"
            else:
                ws["C"+str(length+number+1)] = "その他"
            if re.match(date,fn):
                ws["D"+str(length+number+1)] = "〇"
        length=length+len(list[i])
        
    wb.save("フォルダ内まとめ_update"+kyo_name+".xlsx")
    
    root = Tk()
    root.withdraw()
    messagebox.showinfo("メッセージ", "作業が完了しました")
    root.quit()

if __name__=="__main__":
    makefolder()
